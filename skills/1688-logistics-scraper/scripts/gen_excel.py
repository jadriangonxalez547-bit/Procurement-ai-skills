import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
src = os.path.join(SCRIPT_DIR, 'orders_data.json')
with open(src, 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = '1688待收货物流'

# 表头
headers = ['序号', '采购订单号', '订单状态', '物流名称', '物流单号', '物流更新时间', '最新物流轨迹']
ws.append(headers)

# 表头样式
header_fill = PatternFill('solid', fgColor='FF6B35')
header_font = Font(bold=True, color='FFFFFF', size=11)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin = Side(border_style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col in range(1, len(headers)+1):
    c = ws.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
    c.border = border

# 状态颜色
status_colors = {
    '已签收': 'C6EFCE',     # 绿
    '派送中': 'FFEB9C',     # 黄
    '已揽件': 'BDD7EE',     # 浅蓝
    '运输中': 'FFC7CE',     # 浅红
    '已发货': 'BDD7EE',     # 浅蓝
    '待收货': 'F2F2F2',     # 灰
    '物流异常提醒': 'FF7F50', # 珊瑚色（醒目）
    '无': 'FFFFFF',         # 白
}

# 数据
for i, o in enumerate(data, 1):
    row = [i, o['orderId'], o['status'], o['expressName'], o['expressNo'], o['updateTime'], o['latestTrack']]
    ws.append(row)
    excel_row = i + 1
    fill_color = status_colors.get(o['status'], 'FFFFFF')
    for col in range(1, len(headers)+1):
        c = ws.cell(row=excel_row, column=col)
        c.border = border
        if col == 1 or col == 2 or col == 5 or col == 6:
            c.alignment = center
        else:
            c.alignment = left
        if col == 3:
            c.fill = PatternFill('solid', fgColor=fill_color)
            c.font = Font(bold=True)

# 列宽
col_widths = [6, 22, 10, 16, 22, 20, 60]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 行高
ws.row_dimensions[1].height = 28
for i in range(2, len(data)+2):
    ws.row_dimensions[i].height = 50

# 冻结表头
ws.freeze_panes = 'A2'

# 筛选
ws.auto_filter.ref = f'A1:G{len(data)+1}'

out = os.path.join(SCRIPT_DIR, '1688待收货物流.xlsx')
wb.save(out)
print(f'✅ Excel 已保存: {out}')
print(f'   共 {len(data)} 行数据')
